import customtkinter as ctk
import tkinter as tk
from tkcalendar import DateEntry
from Login import renderSignIn
from openpyxl import load_workbook
from api import confirmation_message
from api import postpone_message
from api import cancellation_message
from api import announcement_message
from openpyxl.utils import get_column_letter
from Login import getCode
import babel.numbers
from tkinter import ttk
from PIL import Image, ImageTk
import time
import datetime
import os
import sys

ctk.set_appearance_mode("dark")

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
details = ""
new_time = ""
btnIndex = 0
leccc_code = getCode()
def getDetails():
    return details

announce_msg= ""

current_directory = os.getcwd()
log_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
log_time = f"[{log_time}]"



def renderHome(code):
    global leccc_code
    leccc_code = code
    Lecturers = {
    "0001":"J.B.HAYFRON",
    "0002":"N.USSIPH",
    "0003":"D.ASAMOAH",
    "0004":"J.K.PANFORD",
    "0005":"F.TWUM",
    "0006":"B.E.OWUSU",
    "0007":"G.ABDUL SALAAM",
    "0008":"B.ARTHUR"
    }

    Courses = {
    "0001":"CSM 388",
    "0002":"CSM 354",
    "0003":"CSM 352",
    "0004":"CSM 374",
    "0005":"CSM 358",
    "0006":"CSM 394",
    "0007":"CSM 376",
    "0008":"CSM 286"
    }

    def postpone_options_window():
        window = ctk.CTkToplevel(homepage)
        window.title("Postpone Class")
        window.geometry("500x650")
        window.attributes('-topmost', True)
        label = ctk.CTkLabel(window,text="Current Free Periods",font=ctk.CTkFont("Arial",size=20,weight="bold"))
        label.pack()

        free_periods = tk.Listbox(window,bg="#191919",font=("Arial",15 ),width=50,fg="white",height=20,highlightthickness=0,selectmode=tk.MULTIPLE)
        free_periods.pack()

        wb=load_workbook(resource_path(f"{current_directory}\\data\\CS3 Schedule.xlsx"))
        sheet = wb[wb.sheetnames[0]]
        wb.active = wb[wb.sheetnames[0]]
        first_col = wb.active.min_column
        last_col = wb.active.max_column
        first_row = wb.active.min_row
        last_row = wb.active.max_row

        
        def is_cell_merged(sheet, row, column):
            # Check if the cell at the given row and column is part of any merged range.
            for merged_range in sheet.merged_cells.ranges:
                if row in range(merged_range.min_row, merged_range.max_row + 1) and column in range(merged_range.min_col, merged_range.max_col + 1):
                    return True
            return False

        period_index = 1
        for i in range(first_row+1,last_row+1):
            for j in range(first_col+1, last_col+1):
                if is_cell_merged(sheet,i,j) and sheet[str(get_column_letter(j))+str(i)].value != None:
                    if sheet[str(get_column_letter(j))+str(i)].value == None:
                        if sheet[str(get_column_letter(1))+str(i)].value != None:
                            detail = f"{sheet[str(get_column_letter(1))+str(i)].value}     -     {sheet[str(get_column_letter(j))+str('1')].value}"
                            free_periods.insert(period_index,detail)
                            period_index +=1
                        else:
                            continue
                elif is_cell_merged(sheet,i,j) and sheet[str(get_column_letter(j))+str(i)].value == None:
                    continue
                else:
                        if sheet[str(get_column_letter(j))+str(i)].value == None:
                            if sheet[str(get_column_letter(1))+str(i)].value != None:
                                detail = f"{sheet[str(get_column_letter(1))+str(i)].value}     -     {sheet[str(get_column_letter(j))+str('1')].value}"
                                free_periods.insert(period_index,detail)
                                period_index +=1

        wb.close()

        def applySelection():
            global new_time
            new_time_list = []
            new_time = ""
            for index in free_periods.curselection():
                new_time_list.insert(index,free_periods.get(index))
            
            for time in new_time_list:
                new_time = new_time + time

            new_time =  merge_time_slots(new_time.replace(' ',''))
        
        def merge_time_slots(input_str):
            time_slots = input_str.split('-')
            merged_slots = []

            day = None
            start_time = None
            end_time = None

            for slot in time_slots:
                if slot.isalpha():
                    if day and start_time and end_time:
                        merged_slots.append(f"{day}-{start_time}-{end_time}")
                        start_time = None
                        end_time = None
                    day = slot
                else:
                    if start_time is None:
                        start_time = slot
                    else:
                        end_time = slot

            if day and start_time and end_time:
                merged_slots.append(f"{day}-{start_time}-{end_time}")

            return '-'.join(merged_slots)


        def close_popup(window):
            window.destroy()


        buttons_frame = ctk.CTkFrame(window)
        buttons_frame.pack()

        select_btn = ctk.CTkButton(buttons_frame,text="Select",command=lambda:(applySelection(),postpone_message(),renderSuccess(),postponeMakeLog(log_time)))
        select_btn.grid(column=0,row = 0,padx = 20)

        cancel_btn = ctk.CTkButton(buttons_frame,text="Cancel", command=lambda:(close_popup(window)))
        cancel_btn.grid(column=1,row = 0)

        success_msg = ctk.CTkLabel(window,text="Success!",font=ctk.CTkFont("Arial",size=12))
        success_msg.pack() 
        success_msg.pack_forget()

        def renderSuccess():
            from api import results
            if results == "Response 200":
                success_msg.configure(text="Success!")
                success_msg.pack()
                window.update()
                time.sleep(2)
                success_msg.configure(text="Logging changes...")
            else:
                success_msg.configure(text="Something went wrong.Please try again or check your internet connection")
                success_msg.pack()
            window.update()
            time.sleep(15)
            success_msg.configure(text="Done")
            window.update()
            time.sleep(1)
            success_msg.pack_forget()
            window.destroy()
     

    def removeIndicators():
        schedule_btn.configure(fg_color = "#2c3032")
        announce_btn.configure(fg_color = "#2c3032")
        history_btn.configure(fg_color = "#2c3032")

    def showIndicator(button):

        removeIndicators()
        button.configure(fg_color ="#5e43f3" )
    

    def renderSchedule():

        destroyFrames()

        wb=load_workbook(resource_path(f"{current_directory}\\data\\CS3 Schedule.xlsx"))
        sheet = wb[wb.sheetnames[0]]
        wb.active = wb[wb.sheetnames[0]]
        first_col = wb.active.min_column
        last_col = wb.active.max_column
        first_row = wb.active.min_row
        last_row = wb.active.max_row

        frame_col = 0
        frame_row = 0

        profile_img= ImageTk.PhotoImage(Image.open(resource_path(f"{current_directory}\\assets\\icons\\profile.png")).resize((49,45)))
        profile_label = ctk.CTkLabel(content_frame,image=profile_img,compound="right",text=f"Welcome, {Lecturers[leccc_code]} ",font=ctk.CTkFont('Terminal',size=15))
        profile_label.pack(side=tk.TOP, anchor=tk.NE) 


        next_class_frame=ctk.CTkFrame(content_frame,width=750,height=70,corner_radius=20,fg_color="#191919")
        next_class_frame.pack()

        ctk.CTkLabel(next_class_frame,text="Your Upcoming Classes",font=ctk.CTkFont("Arial",size=60,weight="bold")).place(relx=0.5,rely=0.5,anchor=ctk.CENTER)

        upcoming_classes_frame = ctk.CTkFrame(content_frame,width=700,height=500)
        upcoming_classes_frame.pack(pady=(60,0))


        tiles= []
        buttonIndex = 0
        confirm_button_id_map = {}
        postpone_button_id_map ={}
        def create_confirm_btn(buttonIndex):
            button = ctk.CTkButton(tile_frame,text="Confirm",font=ctk.CTkFont("Arial",weight="bold"),command=lambda:(onClick(buttonIndex),getDayTime(),confirmation_message(),show_success(),confirmationMakeLog(log_time)))
            confirm_button_id_map[buttonIndex] = button
            button.grid(column =1,row=2)

        def create_postpone_btn(buttonIndex):
            button = ctk.CTkButton(tile_frame,text="Postpone",font=ctk.CTkFont("Arial",weight="bold"),fg_color="#9171f8",hover_color="#ba9ffb",command=lambda:(onClick(buttonIndex),getDayTime(),postpone_options_window()))
            postpone_button_id_map[buttonIndex] = button
            button.grid(column =1,row=3,columnspan=2,sticky="ew")

        def create_cancel_btn(buttonIndex):
            button = ctk.CTkButton(tile_frame,text="Cancel",font=ctk.CTkFont("Arial",weight="bold"),fg_color="#FF007F",hover_color="#FF66B2",command=lambda:(onClick(buttonIndex),getDayTime(),cancellation_message(),show_success(),cancellationMakeLog(log_time)))
            postpone_button_id_map[buttonIndex] = button
            button.grid(column =2,row=2,padx=(10,0))

        def show_success():
            from api import results
            if results == "Response 200":
                success_msg.configure(text="Success!")
                success_msg.pack()
                homepage.update()
                time.sleep(2)
                success_msg.configure(text="Logging Changes...")
            else:
                success_msg.configure(text="Something went wrong.Please try again or check your internet connection")
                success_msg.pack()

            homepage.update()
            time.sleep(3)
            success_msg.pack_forget()


        



        for i in range(first_col+3,last_col+1):
            col = get_column_letter(i)
            for j in range(first_row + 1, last_row + 1):
                if sheet[str(col)+str(j)].value == Lecturers[code]:
                    scheduled_time =  f"{sheet[str(col)+'1'].value},{sheet[str(get_column_letter(i+1))+'1'].value}"
                    scheduled_time = scheduled_time[:4] + scheduled_time[16:]
                    current_day = sheet["A"+str(j)].value
                    if current_day == None:
                        current_day = sheet["A"+str(j-1)].value
                    tile_frame= ctk.CTkFrame(upcoming_classes_frame,corner_radius=10,fg_color="#2c3032",width=910,height=700)
                    tile_frame.grid(column=frame_col,row=frame_row,padx=10,pady=20,sticky="ew")
                    ctk.CTkLabel(tile_frame,text=Courses[code],width=200,height=100,corner_radius=10,fg_color="#404040",font=ctk.CTkFont("Arial",size=20,weight="bold")).grid(column=0,row=0,pady=10,rowspan=3,sticky="ewns",padx=10)
                    ctk.CTkLabel(tile_frame,text=f"{current_day}",font=ctk.CTkFont("Arial",size=25,weight="bold")).grid(column=1,row=0,columnspan=2)
                    ctk.CTkLabel(tile_frame,text=f"{scheduled_time}",font=ctk.CTkFont("Arial",size=20,weight="bold")).grid(column=1,row=1,columnspan=2)
                    create_confirm_btn(buttonIndex)
                    create_postpone_btn(buttonIndex)
                    create_cancel_btn(buttonIndex)
                    tiles.append({
                        "name":Courses[code],
                        "day":current_day,
                        "time":scheduled_time,
                        "id":buttonIndex
                    })
                    frame_col +=1
                    buttonIndex +=1
                    if frame_col == 2:
                        frame_row +=1
                        frame_col = 0
        wb.close()
        success_msg = ctk.CTkLabel(content_frame,text="Success!")
        success_msg.pack()
        success_msg.pack_forget()

        viewSchedule_btn.place(relx=0.44,rely=0.9)

        def getDayTime():
            global details
            global btnIndex
            fullday = ""
            for index, tile in enumerate(tiles):
                if btnIndex == tiles[index]['id']:
                    if tiles[index]['day'] == "MON":
                        fullday = "Monday"
                    elif tiles[index]['day'] == "TUE":
                        fullday = "Tuesday"
                    elif tiles[index]['day'] == "WED":
                        fullday = "Wednesday"
                    elif tiles[index]['day'] == "THU":
                        fullday = "Thursday"                       
                    elif tiles[index]['day'] == "FRI":
                        fullday = "Friday"

                    details = f"{fullday} {tiles[index]['time']} "
            getDetails()
            

        def onClick(buttonIndex):
            global btnIndex
            btnIndex = buttonIndex
            return buttonIndex
        


    def renderAnnounce():
        destroyFrames()

        def show_success():
            from api import results
            if results == "Response 200":
                success_msg.configure(text="Success!  Please wait while we log your announcements...")
                success_msg.place(relx = 0.4 , rely= 0.72, anchor = ctk.CENTER)
            else:

                success_msg.configure(text="Something went wrong.Please try again or check your internet connection")
                success_msg.place(relx = 0.4 , rely= 0.72, anchor = ctk.CENTER)
            homepage.update()
            time.sleep(3)
            success_msg.place_forget()

        ctk.CTkLabel(content_frame,text="Welcome to the Announcements Center",font=ctk.CTkFont("Arial",size=53,weight="bold"),text_color="#5e43f3").pack()
        announce_grid = ctk.CTkFrame(content_frame,fg_color="#191919")
        announce_grid.pack()
        ctk.CTkLabel(announce_grid,text="Communicate with your students",font=ctk.CTkFont("Arial",size=30,weight="normal"),text_color="white").grid(column=0,row=0,padx = 20)
        ctk.CTkLabel(announce_grid,text="Let your voice be heard",font=ctk.CTkFont("Arial",size=30,weight="normal"),text_color="white").grid(column=1,row=0,padx = 20)

        success_msg = ctk.CTkLabel(content_frame,text="Success!  Please wait while we log your announcements...")
        success_msg.place(relx = 0.4 , rely= 0.72, anchor = ctk.CENTER)
        success_msg.place_forget()


        entry_frame = ctk.CTkFrame(content_frame,fg_color="#191919")
        entry_frame.place(relx=0.5,rely=0.85,anchor=ctk.CENTER)

        text_widget = tk.Text(entry_frame, wrap="word", width=60, height=8,background="#404040",foreground="white",borderwidth=0,font=("Arial",16),insertbackground="white")
        text_widget.grid(column=0,row=0,padx=(0,30))

        def getAnnouncement():
            global announce_msg
            announce_msg = text_widget.get("1.0","end-1c")
            text_widget.delete("1.0", tk.END)

        submit_entry_btn = ctk.CTkButton(entry_frame,text="Submit", command=lambda:(getAnnouncement(),announcement_message(),show_success(),announcementMakeLog(log_time)))
        submit_entry_btn.grid(column=1,row=0)
        
    def append_log(log_widget, text):
        log_widget.configure(state=tk.NORMAL)
        log_widget.insert(tk.END, text + "\n")
        log_widget.configure(state=tk.DISABLED)

    def updateHistory(log_text):
        log_text.configure(state=tk.NORMAL)
        log_text.delete(1.0, tk.END)
        log_text.configure(state="disabled") 

    def renderHistory():
        from docs import document_id,read_google_doc
        from filter import extract_logs,main
        global text_box
        destroyFrames()
        homepage.update()
        time.sleep(1)

        
        def filterHistory():
            start_date = start_calendar.get_date()
            end_date = end_calendar.get_date()
            start_date = start_date.strftime("%Y-%m-%d")
            end_date = end_date.strftime("%Y-%m-%d")
            filter_action= action_combo.get()
            filtered_logs = extract_logs(main(),start_date,end_date,filter_action)
            updateHistory(history_box)
            for log in filtered_logs:
                append_log(history_box,log)

        def on_entry_click(event,calendar):
            if calendar.get() == 'Start Date':
                calendar.delete(0, tk.END)
                calendar.configure(foreground='black')

        def on_focus_out(event,calendar):
            if calendar.get() == '':
                calendar.insert(0, 'Start Date')
                calendar.configure(foreground='gray')


        filter_frame = ttk.LabelFrame(content_frame,text="Filter")
        filter_frame.pack()
        action_comboList = ["All Actions","Confirm","Cancel","Postpone","Announcement"]

        start_calendar = DateEntry(filter_frame, width=20, background='#2c3032',foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_calendar.delete(0,tk.END)
        start_calendar.insert(0, 'Start date')
        start_calendar.bind('<FocusIn>', on_entry_click(on_entry_click,start_calendar))
        start_calendar.bind('<FocusOut>', on_focus_out(on_focus_out,start_calendar))
        start_calendar.configure(foreground='gray')
        start_calendar.grid(column=0,row=0,padx=10)

        end_calendar = DateEntry(filter_frame, width=20, background='darkblue',foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_calendar.delete(0,tk.END)
        end_calendar.insert(0, 'End date')
        end_calendar.bind('<FocusIn>', on_entry_click(on_entry_click,end_calendar))
        end_calendar.bind('<FocusOut>', on_focus_out(on_focus_out,end_calendar))
        end_calendar.configure(foreground='gray')
        end_calendar.grid(column=1,row=0)   

        action_combo = ctk.CTkComboBox(filter_frame,values=action_comboList)
        action_combo.grid(column=2,row=0,padx=10)

        apply_filter_btn= ctk.CTkButton(filter_frame,text="Apply",command=filterHistory)
        apply_filter_btn.grid(column=3,row=0)

        fullHistory_btn= ctk.CTkButton(filter_frame,text="View Full history",command=lambda:(updateHistory(history_box),append_log(history_box, read_google_doc(document_id)),action_combo.set("All Actions")))
        fullHistory_btn.grid(column=4,row=0,padx=10,pady=10)

        
        history_box = ctk.CTkTextbox(content_frame,height=500,width=750,wrap = ctk.WORD,state=ctk.DISABLED)
        history_box.pack()

        title_label = ctk.CTkLabel(content_frame,text="LecCheck History",font=ctk.CTkFont("Arial",size=53,weight="bold"),text_color="white")
        title_label.pack()
        subhead1_text = ctk.CTkLabel(content_frame,text="Stay up to date with events and changes",font=ctk.CTkFont("Arial",size=25),text_color="#5e43f3").pack()
        subhead2_text = ctk.CTkLabel(content_frame,text="Never miss out",font=ctk.CTkFont("Arial",size=25),text_color="white").pack()
        append_log(history_box, read_google_doc(document_id))



    def destroyFrames():
        for frame in content_frame.winfo_children():
            frame.forget()
            # frame.pack_forget()
            frame.grid_forget()
            frame.place_forget()
            viewSchedule_btn.place_forget()


    def refreshHistory():
        renderHistory()
        homepage.update()

    def announcementMakeLog(logTime):
        from docs import document_id,append_text_to_google_doc
        text = f"{logTime}: {Lecturers[leccc_code]}: ANNOUNCEMENT!!!! {announce_msg}"
        append_text_to_google_doc(document_id,text)

    def confirmationMakeLog(logTime):
        from docs import document_id,append_text_to_google_doc
        text = f"{logTime}: {Lecturers[leccc_code]}: Please be informed. The {Courses[leccc_code]} lecture for {getDetails()}has been CONFIRMED. See you in class! "
        append_text_to_google_doc(document_id,text)

    def cancellationMakeLog(logTime):
        from docs import document_id,append_text_to_google_doc
        text = f"{logTime}: {Lecturers[leccc_code]}: Please be informed. The {Courses[leccc_code]} lecture for {getDetails()}has been CANCELED. Please stay tuned for any more announcements"
        append_text_to_google_doc(document_id,text)

    def postponeMakeLog(logTime):
        from docs import document_id,append_text_to_google_doc
        text = f"{logTime}: {Lecturers[leccc_code]}: Please be informed. The {Courses[leccc_code]} lecture for {getDetails()}has been POSTPONED. The new meeting time is: {new_time}"
        append_text_to_google_doc(document_id,text)


    def show_overlay():
        if overlay_frame.winfo_ismapped():
            hide_overlay()
        else:
            overlay_frame.place(relx=0.013, rely=0.7)
            overlay_frame.lift()

    def hide_overlay():
        overlay_frame.place_forget()

    def backToSignIn():
        homepage.destroy()
        from Login import renderSignIn
        renderSignIn()

    def openSchedule():
        os.startfile(resource_path(f"{current_directory}\\data\\CS3 Schedule.xlsx"))

    def on_map(event):
        homepage.state("zoomed")

    homepage = ctk.CTk()
    homepage.geometry("1250x750")
    homepage.bind("<Map>", on_map)
    homepage.title("LecCheck")
    
    nav_pane = ctk.CTkFrame(homepage,corner_radius=0,fg_color="#2c3032")
    nav_pane.pack_propagate(False)
    nav_pane.pack(side=ctk.LEFT,fill="y")

    content_frame = ctk.CTkFrame(homepage,corner_radius=0, fg_color="#191919")
    content_frame.pack_propagate(False)
    content_frame.pack(side=ctk.LEFT, fill="both",expand=True)

    logo_img = ImageTk.PhotoImage(Image.open(resource_path(f"{current_directory}\\assets\\icons\\logo.png")).resize((50,50)))
    logoTitle = ctk.CTkLabel(nav_pane,height=77,corner_radius=0,text="LecCheck",font=ctk.CTkFont("Arial",weight="bold",size=15),image=logo_img, compound="left")
    logoTitle.grid(column=0,row=0)

    settings_img= ImageTk.PhotoImage(Image.open(resource_path(f"{current_directory}\\assets\\icons\\settings.png")).resize((40,40)))
    settings_button = ctk.CTkButton(nav_pane,image=settings_img,width=20,text="",fg_color="#2c3032",corner_radius=0,hover_color="#2c3032",command=show_overlay)
    settings_button.place(relx=0.7, rely=1, anchor="sw")

    overlay_frame = ctk.CTkFrame(content_frame,fg_color="#2c3032",height=150,width=130,corner_radius=5)
    overlay_frame.pack_propagate(False)
    overlay_frame.place(relx=0.013, rely=0.7)
    overlay_frame.place_forget()

    signout_btn = ctk.CTkButton(overlay_frame,text="Sign Out",command=backToSignIn)
    signout_btn.grid(column=0,row=0,pady=6,padx=5)

    refresh_btn = ctk.CTkButton(overlay_frame,text="Refresh",command=refreshHistory)
    refresh_btn.grid(column=0,row=1,pady=6,padx=5)

    profile_btn = ctk.CTkButton(overlay_frame,text="Profile")
    profile_btn.grid(column=0,row=2,pady=6,padx=5)

    schedule_img = ImageTk.PhotoImage(Image.open(resource_path(f"{current_directory}\\assets\\icons\\calendar.png")).resize((50,50)))
    schedule_btn=ctk.CTkButton(nav_pane,text="Schedule",anchor=ctk.W,image=schedule_img,compound="left",corner_radius=0,fg_color="#2c3032",font=ctk.CTkFont("Arial",weight="bold"),command=lambda:(showIndicator(schedule_btn),renderSchedule()),hover_color="#5e43f3")
    schedule_btn.grid(column=0,row=1,sticky="nsew")


    announce_img=ImageTk.PhotoImage(Image.open(resource_path(f"{current_directory}\\assets\\icons\\megaphone.png")).resize((50,50)))
    announce_btn = ctk.CTkButton(nav_pane,text="Announcements",anchor=ctk.W,image=announce_img,compound="left",corner_radius=0,fg_color="#2c3032",font=ctk.CTkFont("Arial",weight="bold"),command=lambda:(showIndicator(announce_btn),renderAnnounce()),hover_color="#5e43f3")
    announce_btn.grid(column=0,row=3, sticky="nsew")

    history_img=ImageTk.PhotoImage(Image.open(resource_path(f"{current_directory}\\assets\\icons\\history1.png")).resize((53,53)))
    history_btn = ctk.CTkButton(nav_pane,text="History",anchor=ctk.W,image=history_img,compound="left",corner_radius=0,fg_color="#2c3032",font=ctk.CTkFont("Arial",weight="bold"),command=lambda:(showIndicator(history_btn),renderHistory()),hover_color="#5e43f3")
    history_btn.grid(column=0,row=4, sticky="nsew")

    viewSchedule_btn = ctk.CTkButton(content_frame,text="View Full Schedule",command=openSchedule)
    viewSchedule_btn.place(relx=0.44,rely=0.9)
    viewSchedule_btn.place_forget()


    showIndicator(schedule_btn)
    renderSchedule()
    
    homepage.tk.call("source", resource_path("assets\\Azure-ttk-theme-main\\azure.tcl"))
    homepage.tk.call("set_theme", "dark")

    homepage.mainloop()